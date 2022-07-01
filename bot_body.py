import openpyxl
import telebot

file = 'black_list.xlsx'
token = 
bot = telebot.TeleBot(token)


#Приведение номера к формату. 1, если номер некорректен
def correct_number(number):
    a = [' ', '+', '(', ')', '-']
    for s in a:
        number = number.replace(s, '')
    if number.startswith(('7', '8')) and len(number) == 11 and number.isdigit():
        number = f'+7 ({number[1:4]}) {number[4:7]}-{number[7:9]}-{number[9:]}'
        return number
    elif number.startswith(('9', '8')) and len(number) == 10 and number.isdigit():
        number = '1' + number
        number = f'+7 ({number[1:4]}) {number[4:7]}-{number[7:9]}-{number[9:]}'
        return number
    else:
        return 1


#Ввод номера
def start_list(message):
    mg = bot.send_message(message.chat.id, 'Введите номер телефона')
    bot.register_next_step_handler(mg, phon_number)


#Проверка номера, Ввод адреса
def phon_number(message):
    phon = message.text
    phon = correct_number(phon)
    if phon == 1:
        bot.send_message(message.chat.id, 'Номер не существует. Попробуйте еще раз.')
        start_list(message)

    else:
        client_info = [phon]
        address = bot.send_message(message.chat.id, 'Введите адрес')
        bot.register_next_step_handler(address, address_phone, client_info)


#Ввод проблемы
def address_phone(message, client_info):
    client_info.append(message.text)
    description = bot.send_message(message.chat.id, 'Опишите происшествие')
    bot.register_next_step_handler(description, address_phone_description, client_info)


#Добавление проблемы в инфо
def address_phone_description(message, client_info):
    client_info.append(message.text)
    data_writer(client_info, file)
    bot.send_message(message.chat.id, 'Запись добавлена')


#Запись в файл
def data_writer(client_info, output_file):
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook['data']
    ws.append(client_info)
    workbook.save(output_file)
    workbook.close()


#инициализация проверки номера, поиска по номеру(если существует) и вывод инфы
def phon_number_search(message):
    phon = message.text
    phon = correct_number(phon)
    if phon == 1:
        bot.send_message(message.chat.id, 'Номер не существует. Попробуйте еще раз.')
        client_search_by_number(message)

    else:
        clients_info = '\n'.join(data_reader(phon, file))
        bot.send_message(message.chat.id, clients_info)


#инициализация поиска по адресу и вывод
def address_search(message):
    address = message.text
    clients_info = '\n'.join(data_reader_address(address, file))
    bot.send_message(message.chat.id, clients_info)



#поиск инфы о клиенте по адресу(введенному вручную)
def data_reader_address(address, output_file):
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook['data']
    max_row = 'C' + str(ws.max_row)
    client_list = []
    for row in ws['A2':max_row]:
        if address in row[1].value:
            client_list.append(f'{row[0].value}\n{row[1].value}\n{row[2].value}')
    workbook.close()
    if len(client_list) == 0:
        return ['Клиент чист']
    elif len(client_list) == 1:
        return client_list
    else:
        return [f'{i+1}: {client}' for i, client in enumerate(client_list)]


#поиск инфы о клиенте по номеру.
def data_reader(client_number, output_file):
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook['data']
    max_row = 'C' + str(ws.max_row)
    client_list = []
    for row in ws['A2':max_row]:
        if client_number == row[0].value:
            client_list.append(f'{row[0].value}\n{row[1].value}\n{row[2].value}')
    workbook.close()
    if len(client_list) == 0:
        return ['Клиент чист']
    elif len(client_list) == 1:
        return client_list
    else:
        return [f'{i+1}: {client}' for i, client in enumerate(client_list)]


#Внесение данных
@bot.message_handler(commands=['start'])
def start_add(message):
    start_list(message)


#поиск по параметру. вызывает пачку кнопок. зачем? загадка...
@bot.message_handler(commands=['search_number'])
def client_search_by_number(message):
    mg = bot.send_message(message.chat.id, 'Введите номер телефона')
    bot.register_next_step_handler(mg, phon_number_search)


@bot.message_handler(commands=['search_address'])
def client_search_by_address(message):
    mg = bot.send_message(message.chat.id, 'Введите адрес(без номера квартиры)')
    bot.register_next_step_handler(mg, address_search)


bot.polling(none_stop=True)
