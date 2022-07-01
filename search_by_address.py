import openpyxl
import telebot


file = 'black_list.xlsx'
token = 
bot = telebot.TeleBot(token)


#Функция должна вернуть адрес определенного формата(хз какого) и 1(цифру 1), если адресс некорректен
def correct_address(address):
    pass


@bot.message_handler(commands=['search_address'])
def client_search_address(message):
    mg = bot.send_message(message.chat.id, 'Введите номер телефона')
    bot.register_next_step_handler(mg, address_search)


def address_search(message):
    address = message.text
    #address = correct_address(address)
    if address == 1:
        bot.send_message(message.chat.id, 'Адресс некорректен. Попробуйте еще раз.')
        client_search_address(message)

    else:
        clients_info = '\n'.join(data_reader(address, file))
        bot.send_message(message.chat.id, clients_info)


def data_reader(client_number, output_file):
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook['data']
    max_row = 'C' + str(ws.max_row)
    client_list = []
    for row in ws['A2':max_row]:
        if client_number in row[1].value:          #Тут надо что-то придумать со сравением адресов
            client_list.append(f'{row[1].value}\n{row[2].value}\n{row[0].value}')
    workbook.close()
    if len(client_list) == 0:
        return ['Клиент чист']
    elif len(client_list) == 1:
        return client_list
    else:
        return [f'{i+1}: {client}' for i, client in enumerate(client_list)]


bot.polling(none_stop=True)
