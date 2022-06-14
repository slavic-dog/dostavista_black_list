import openpyxl
import telebot


file = 'black_list.xlsx'
token = '5344175403:AAE2sQOhTKvZLhxji_W_LwHwr-P0Jhv4gfU'
bot = telebot.TeleBot(token)


def correct_number(number):
    a = [' ', '+', '(', ')', '-']
    for s in a:
        number = number.replace(s, '')
    if number.startswith(('7', '8')) and len(number) == 11 and number.isdigit():
        number = f'+7 ({number[1:4]}) {number[4:7]}-{number[7:9]}-{number[9:]}'
        return number
    elif number.startswith(('9', '8')) and len(number) == 10 and number.isdigit():
        number = '1' + number
        number = f'+7 ({number[1:4]}) {number[4:7]}-{number[7:9]}-{number[9:]}'
        return number
    else:
        return 1


@bot.message_handler(commands=['search_number'])
def client_search(message):
    mg = bot.send_message(message.chat.id, 'Введите номер телефона')
    bot.register_next_step_handler(mg, phon_number_search)


def phon_number_search(message):
    phon = message.text
    phon = correct_number(phon)
    if phon == 1:
        bot.send_message(message.chat.id, 'Номер не существует. Попробуйте еще раз.')
        client_search(message)

    else:
        clients_info = '\n'.join(data_reader(phon, file))
        bot.send_message(message.chat.id, clients_info)


def data_reader(client_number, output_file):
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook['data']
    max_row = 'C' + str(ws.max_row)
    client_list = []
    for row in ws['A2':max_row]:
        if client_number == row[0].value:
            client_list.append(f'{row[2].value}\n{row[1].value}')
    workbook.close()
    if len(client_list) == 0:
        return ['Клиент чист']
    elif len(client_list) == 1:
        return client_list
    else:
        return [f'{i+1}: {client}' for i, client in enumerate(client_list)]


bot.polling(none_stop=True)
